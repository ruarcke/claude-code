#!/usr/bin/env python3
"""
Gera planilha RCPCC direto do Kommo.
Lê o "Resumo do caso", usa Claude para extrair dados, preenche o template Excel.

Uso: python gerar_planilha_kommo.py <lead_id>
"""

import sys
import os
import json
import shutil
import re

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import requests
except ImportError:
    print("ERRO: requests nao instalado. Execute: pip install requests")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip install openpyxl")
    sys.exit(1)

# Config - chaves importadas de config.py (não commitado no git)
from config import KOMMO_TOKEN, CLAUDE_KEY
TEMPLATE_PATH = os.path.join(os.path.expanduser('~'), 'Desktop', 'Tabela atendimento - RCPCC.xlsx')
OUTPUT_FOLDER = os.path.join(os.path.expanduser('~'), 'Desktop', 'ATENDIMENTOS', 'RCPCC - Atendimentos')


def buscar_lead(lead_id):
    """Busca lead no Kommo e extrai Resumo do caso."""
    print(f"📡 Buscando lead {lead_id} no Kommo...")
    resp = requests.get(
        f"https://ruarcke.kommo.com/api/v4/leads/{lead_id}",
        headers={"Authorization": KOMMO_TOKEN, "Accept": "application/json"},
        timeout=30
    )
    resp.raise_for_status()
    lead = resp.json()

    nome = lead.get('name', 'Cliente')
    resumo = ''
    for field in lead.get('custom_fields_values', []):
        if field.get('field_name', '').lower().startswith('resumo'):
            resumo = field['values'][0]['value']
            break

    if not resumo:
        print("ERRO: Campo 'Resumo do caso' vazio ou nao encontrado no card.")
        sys.exit(1)

    print(f"👤 Lead: {nome}")
    print(f"📋 Resumo encontrado ({len(resumo)} chars)")
    return nome, resumo


def extrair_dividas(resumo):
    """Usa Claude para extrair dados estruturados do resumo."""
    print("🤖 Extraindo dados das dividas via Claude...")

    prompt = f"""Extraia os dados de dívidas do texto abaixo e retorne APENAS um JSON válido (sem markdown, sem código, apenas o JSON puro).

Estrutura esperada:
{{
  "dividas": [
    {{
      "banco": "Nome do Banco",
      "modalidade": "Cartão/Empréstimo/Cheque Especial/Renegociação",
      "valor_parcela": 0,
      "quantidade": 1,
      "pagas": 0,
      "em_atraso": 1
    }}
  ],
  "tem_processo": false,
  "observacoes": ""
}}

Regras IMPORTANTES:
- Os valores NUNCA podem ser 0 se o texto mencionar algum valor. Se diz "R$ 112.000" e menciona cartão e cheque especial, distribua proporcionalmente ou coloque tudo em uma linha.
- valor_parcela = valor total da divida naquela modalidade
- quantidade = numero de parcelas (se nao especificado, use 1)
- pagas = parcelas pagas (se nao especificado, use 0)
- em_atraso = quantidade - pagas
- Cartão e cheque especial no mesmo banco: separe em duas linhas
- Sem modalidade especificada: use "Empréstimo"
- Valores devem ser numéricos (sem R$, sem pontos de milhar). Ex: 112000
- Coloque observações relevantes

Texto do caso:
{resumo}"""

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": CLAUDE_KEY,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json"
        },
        json={
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=60
    )
    resp.raise_for_status()
    text = resp.json()['content'][0]['text']

    # Parse JSON
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    data = json.loads(text.strip())

    dividas = data.get('dividas', [])
    print(f"✅ {len(dividas)} divida(s) encontrada(s)")
    for d in dividas:
        print(f"   • {d['banco']} - {d['modalidade']}: R$ {d['valor_parcela']:,.0f}".replace(',', '.'))

    return data


def preencher_template(nome, dados, output_path):
    """Preenche o template Excel com os dados extraidos."""
    print("📝 Preenchendo template Excel...")

    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERRO: Template nao encontrado: {TEMPLATE_PATH}")
        sys.exit(1)

    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    dividas = dados.get('dividas', [])
    row_idx = 2
    for d in dividas:
        if row_idx > 14:
            break
        ws[f'A{row_idx}'] = d.get('banco', '')
        ws[f'B{row_idx}'] = d.get('modalidade', '')
        ws[f'C{row_idx}'] = float(d.get('valor_parcela', 0))
        ws[f'D{row_idx}'] = int(d.get('quantidade', 1))
        ws[f'E{row_idx}'] = int(d.get('pagas', 0))
        # F (em atraso) e G (total) tem formulas - nao sobrescrever
        row_idx += 1

    wb.save(output_path)
    wb.close()
    print(f"✅ Planilha salva localmente: {output_path}")
    return output_path


def upload_kommo(lead_id, file_path):
    """Faz upload da planilha para o campo 'documentos' do card no Kommo."""
    DOCS_FIELD_ID = 2080560
    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)

    print(f"📤 Fazendo upload para o Kommo ({file_size} bytes)...")

    # Step 1: Criar sessao de upload
    sess = requests.post('https://drive-c.kommo.com/v1.0/sessions',
        headers={'Authorization': KOMMO_TOKEN, 'Content-Type': 'application/json'},
        json={'file_name': file_name, 'file_size': file_size,
              'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
        timeout=30
    )
    sess.raise_for_status()
    sess_data = sess.json()

    # Step 2: Upload do arquivo
    with open(file_path, 'rb') as f:
        file_data = f.read()
    upload_resp = requests.post(sess_data['upload_url'],
        headers={'Authorization': KOMMO_TOKEN, 'Content-Type': 'application/octet-stream'},
        data=file_data, timeout=60
    )
    upload_resp.raise_for_status()
    upload_result = upload_resp.json()
    file_uuid = upload_result['uuid']
    version_uuid = upload_result['version_uuid']

    # Step 3: Anexar no campo 'documentos' do card
    update_resp = requests.patch(
        f'https://ruarcke.kommo.com/api/v4/leads/{lead_id}',
        headers={'Authorization': KOMMO_TOKEN, 'Content-Type': 'application/json'},
        json={'custom_fields_values': [{
            'field_id': DOCS_FIELD_ID,
            'values': [{'value': {
                'file_uuid': file_uuid,
                'version_uuid': version_uuid,
                'file_name': file_name
            }}]
        }]},
        timeout=30
    )
    update_resp.raise_for_status()
    print(f"✅ Planilha anexada no campo 'documentos' do card!")
    return file_uuid


def main():
    if len(sys.argv) < 2:
        lead_id = input("Digite o ID do lead no Kommo: ").strip()
    else:
        lead_id = sys.argv[1]

    if not lead_id.isdigit():
        print("ERRO: ID do lead deve ser numerico")
        sys.exit(1)

    # 1. Buscar lead no Kommo
    nome, resumo = buscar_lead(lead_id)

    # 2. Extrair dados via Claude
    dados = extrair_dividas(resumo)

    # 3. Preencher template
    nome_limpo = nome.replace('*', '').strip()
    output_name = f"Tabela atendimento - RCPCC - {nome_limpo}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)

    preencher_template(nome_limpo, dados, output_path)

    # 4. Upload para Kommo
    upload_kommo(lead_id, output_path)

    print()
    print("=" * 50)
    print(f"Planilha gerada para: {nome_limpo}")
    print(f"Local: {output_path}")
    print(f"Kommo: Disponivel no campo 'documentos' do card")
    print()
    print("Proximo passo:")
    print("1. Baixe do Kommo ou abra do local")
    print("2. Confira os valores na reuniao")
    print("3. Salve na pasta e execute gerar_apresentacao.bat")
    print("=" * 50)


if __name__ == '__main__':
    main()
